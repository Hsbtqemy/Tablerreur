"""Tests for exporters."""

from __future__ import annotations

import csv
import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from spreadsheet_qa.core.exporters import (
    AnnotatedXLSXExporter,
    CSVExporter,
    IssuesCSVExporter,
    TXTReporter,
    XLSXExporter,
    build_annotated_dataframe,
)
from spreadsheet_qa.core.models import Issue, IssueStatus, Severity


@pytest.fixture
def sample_df():
    return pd.DataFrame(
        {
            "Titre": ["Introduction ; à la vie", '"Quoted"', "Normal"],
            "Auteur": ["Jean Dupont", "Marie Martin", "Pierre"],
            "Date": ["2021", "2020", "2019"],
        }
    )


@pytest.fixture
def sample_issues():
    return [
        Issue.create(
            rule_id="generic.hygiene.leading_trailing_space",
            severity=Severity.WARNING,
            row=0,
            col="Titre",
            original="Introduction ; à la vie ",
            message="Trailing space",
            suggestion="Introduction ; à la vie",
        )
    ]


class TestCSVExporter:
    def test_delimiter_is_semicolon(self, sample_df, tmp_path):
        path = tmp_path / "out.csv"
        CSVExporter().export(sample_df, path)
        content = path.read_text(encoding="utf-8")
        # Parse with semicolon — should produce correct columns
        reader = csv.reader(io.StringIO(content), delimiter=";")
        rows = list(reader)
        assert rows[0] == ["Titre", "Auteur", "Date"]

    def test_values_with_semicolon_are_quoted(self, sample_df, tmp_path):
        path = tmp_path / "out.csv"
        CSVExporter().export(sample_df, path)
        content = path.read_text(encoding="utf-8")
        # "Introduction ; à la vie" contains ; → must be quoted
        assert '"Introduction ; à la vie"' in content

    def test_utf8_no_bom_by_default(self, sample_df, tmp_path):
        path = tmp_path / "out.csv"
        CSVExporter().export(sample_df, path, bom=False)
        raw = path.read_bytes()
        assert not raw.startswith(b"\xef\xbb\xbf")  # no BOM

    def test_utf8_bom_when_requested(self, sample_df, tmp_path):
        path = tmp_path / "out.csv"
        CSVExporter().export(sample_df, path, bom=True)
        raw = path.read_bytes()
        assert raw.startswith(b"\xef\xbb\xbf")

    def test_roundtrip(self, sample_df, tmp_path):
        path = tmp_path / "out.csv"
        CSVExporter().export(sample_df, path)
        df2 = pd.read_csv(path, sep=";", dtype=str)
        assert list(df2.columns) == list(sample_df.columns)
        assert len(df2) == len(sample_df)


class TestXLSXExporter:
    def test_creates_xlsx_file(self, sample_df, tmp_path):
        path = tmp_path / "out.xlsx"
        XLSXExporter().export(sample_df, path)
        assert path.exists()

    def test_xlsx_has_correct_data(self, sample_df, tmp_path):
        path = tmp_path / "out.xlsx"
        XLSXExporter().export(sample_df, path)
        df2 = pd.read_excel(path, dtype=str, engine="openpyxl")
        assert list(df2.columns) == list(sample_df.columns)
        assert len(df2) == len(sample_df)


class TestAnnotatedExports:
    def test_build_annotated_dataframe_adds_annotation_columns(self, sample_df, sample_issues):
        annotated = build_annotated_dataframe(sample_df, sample_issues, row_positions=[0, 2])
        assert annotated.columns[0] == "__tablerreur_ligne"
        assert "__tablerreur_statut" in annotated.columns
        assert "__tablerreur_anomalies" in annotated.columns
        assert annotated.iloc[0]["__tablerreur_ligne"] == 1
        assert annotated.iloc[0]["__tablerreur_statut"] == "WARNING"
        assert "Titre: Trailing space" in annotated.iloc[0]["__tablerreur_anomalies"]
        assert annotated.iloc[1]["__tablerreur_statut"] == ""

    def test_annotated_xlsx_marks_issue_and_touched_cells(self, sample_df, sample_issues, tmp_path):
        path = tmp_path / "annotated.xlsx"
        AnnotatedXLSXExporter().export(
            sample_df,
            path,
            sample_issues,
            touched_cells={(1, "Auteur")},
            include_visual_marks=True,
            include_status_column=True,
        )

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        issue_cell = ws["B2"]  # __tablerreur_ligne is inserted in column A
        touched_cell = ws["C3"]
        status_cell = ws["E2"]

        assert issue_cell.comment is not None
        assert "Trailing space" in issue_cell.comment.text
        assert issue_cell.fill.fill_type == "solid"
        assert issue_cell.fill.fgColor.rgb.endswith("FDF0D5")
        assert touched_cell.fill.fill_type == "solid"
        assert touched_cell.fill.fgColor.rgb.endswith("DCEFFE")
        assert status_cell.fill.fill_type == "solid"


class TestIssuesCSVExporter:
    def test_delimiter_is_semicolon(self, sample_issues, tmp_path):
        path = tmp_path / "issues.csv"
        IssuesCSVExporter().export(sample_issues, path)
        content = path.read_text(encoding="utf-8")
        reader = csv.reader(io.StringIO(content), delimiter=";")
        rows = list(reader)
        # First row = French headers
        assert "identifiant" in rows[0]
        assert "sévérité" in rows[0]
        assert "colonne" in rows[0]

    def test_row_is_1_based(self, sample_issues, tmp_path):
        path = tmp_path / "issues.csv"
        IssuesCSVExporter().export(sample_issues, path)
        content = path.read_text(encoding="utf-8")
        reader = csv.DictReader(io.StringIO(content), delimiter=";")
        rows = list(reader)
        assert rows[0]["ligne"] == "1"  # French header; 0 → 1 in export


class TestTXTReporter:
    def test_creates_report_file(self, sample_issues, tmp_path):
        path = tmp_path / "report.txt"
        TXTReporter().export(sample_issues, path)
        assert path.exists()

    def test_report_contains_severity(self, sample_issues, tmp_path):
        path = tmp_path / "report.txt"
        TXTReporter().export(sample_issues, path)
        content = path.read_text(encoding="utf-8")
        # Report uses French severity labels
        assert "Avertissement" in content

    def test_report_contains_column_name(self, sample_issues, tmp_path):
        path = tmp_path / "report.txt"
        TXTReporter().export(sample_issues, path)
        content = path.read_text(encoding="utf-8")
        assert "Titre" in content

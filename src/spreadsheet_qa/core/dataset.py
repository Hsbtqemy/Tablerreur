"""DatasetLoader: import CSV and XLSX files into a pandas DataFrame.

Handles:
- Encoding detection via chardet (first 32 KB)
- Delimiter detection via csv.Sniffer (fallback: ; then ,)
- Header row selection (0-based index; rows above it are skipped)
- Sheet selection for XLSX
- Returns (DataFrame, DatasetMeta)
"""

from __future__ import annotations

import codecs
import csv
import hashlib
import io
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import chardet
import pandas as pd

from spreadsheet_qa.core.models import DatasetMeta


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class DatasetLoader:
    """Load CSV or XLSX files into a pandas DataFrame."""

    # Candidate delimiters tried in order when Sniffer fails
    _FALLBACK_DELIMITERS = [";", ",", "\t", "|"]

    def load(
        self,
        path: str | Path,
        header_row: int = 0,
        sheet_name: str | int | None = 0,
        encoding_hint: str | None = None,
        delimiter_hint: str | None = None,
    ) -> tuple[pd.DataFrame, DatasetMeta]:
        """Load a file and return (DataFrame, DatasetMeta).

        Args:
            path: Absolute path to the CSV or XLSX file.
            header_row: 0-based index of the row to use as column headers.
                        Rows before it are skipped entirely.
            sheet_name: For XLSX: sheet name or 0-based integer index.
            encoding_hint: Override encoding detection.
            delimiter_hint: Override delimiter detection (CSV only).

        Returns:
            A tuple of (DataFrame with string dtype, DatasetMeta).
        """
        path = Path(path)
        suffix = path.suffix.lower()

        raw_bytes = path.read_bytes()
        fingerprint = hashlib.sha256(raw_bytes[:65536]).hexdigest()

        if suffix in {".xlsx", ".xls", ".xlsm", ".ods"}:
            df, meta = self._load_xlsx(path, raw_bytes, fingerprint, header_row, sheet_name)
        elif suffix in {".csv", ".tsv", ".txt"}:
            df, meta = self._load_csv(
                path, raw_bytes, fingerprint, header_row, encoding_hint, delimiter_hint
            )
        else:
            # Try CSV as fallback
            df, meta = self._load_csv(
                path, raw_bytes, fingerprint, header_row, encoding_hint, delimiter_hint
            )

        return df, meta

    # ------------------------------------------------------------------
    # XLSX
    # ------------------------------------------------------------------

    def _load_xlsx(
        self,
        path: Path,
        raw_bytes: bytes,
        fingerprint: str,
        header_row: int,
        sheet_name: str | int | None,
    ) -> tuple[pd.DataFrame, DatasetMeta]:
        # Open once via ExcelFile so we can resolve the sheet name without a
        # second openpyxl.load_workbook() call.
        with pd.ExcelFile(path, engine="openpyxl") as xf:
            all_sheets = xf.sheet_names  # list[str], always available

            # Resolve the target sheet selector → canonical name
            if isinstance(sheet_name, str):
                resolved_sheet = sheet_name
            else:
                idx = sheet_name if isinstance(sheet_name, int) else 0
                resolved_sheet = all_sheets[idx] if all_sheets else "Sheet1"

            df_raw = xf.parse(
                resolved_sheet,
                header=None,
                dtype=str,
            )

        df = self._apply_header_row(df_raw, header_row)
        shape = (len(df), len(df.columns))

        meta = DatasetMeta(
            file_path=str(path.resolve()),
            encoding="utf-8",
            delimiter=None,
            sheet_name=resolved_sheet,
            header_row=header_row,
            skip_rows=header_row,
            original_shape=shape,
            column_order=list(df.columns),
            fingerprint=fingerprint,
        )
        return df, meta

    # ------------------------------------------------------------------
    # CSV
    # ------------------------------------------------------------------

    def _load_csv(
        self,
        path: Path,
        raw_bytes: bytes,
        fingerprint: str,
        header_row: int,
        encoding_hint: str | None,
        delimiter_hint: str | None,
    ) -> tuple[pd.DataFrame, DatasetMeta]:
        encoding = encoding_hint or self._detect_encoding(raw_bytes)
        delimiter = delimiter_hint or self._detect_delimiter(raw_bytes, encoding)

        # Use csv.reader directly so ragged rows (e.g. a metadata line that uses
        # a different delimiter) don't cause pandas to infer the wrong column count
        # and drop subsequent rows.
        all_rows = self._read_csv_raw(path, delimiter, encoding)
        if not all_rows:
            df_raw = pd.DataFrame(dtype=str)
        else:
            max_cols = max(len(r) for r in all_rows)
            # Pad short rows (e.g. metadata lines) to uniform width
            padded = [r + [""] * (max_cols - len(r)) for r in all_rows]
            df_raw = pd.DataFrame(padded, dtype=str)

        df = self._apply_header_row(df_raw, header_row)
        shape = (len(df), len(df.columns))

        meta = DatasetMeta(
            file_path=str(path.resolve()),
            encoding=encoding,
            delimiter=delimiter,
            sheet_name=None,
            header_row=header_row,
            skip_rows=header_row,
            original_shape=shape,
            column_order=list(df.columns),
            fingerprint=fingerprint,
        )
        return df, meta

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _apply_header_row(df_raw: pd.DataFrame, header_row: int) -> pd.DataFrame:
        """Use row at `header_row` as column names, drop rows up to and including it."""
        if header_row >= len(df_raw):
            raise ValueError(
                f"header_row={header_row} is beyond the file length ({len(df_raw)} rows)"
            )

        # Extract header values from the target row
        header_values = df_raw.iloc[header_row].tolist()
        # Deduplicate column names (pandas-style)
        seen: dict[str, int] = {}
        deduped: list[str] = []
        for val in header_values:
            name = str(val) if pd.notna(val) and str(val).strip() else f"Unnamed_{len(deduped)}"
            if name in seen:
                seen[name] += 1
                name = f"{name}.{seen[name]}"
            else:
                seen[name] = 0
            deduped.append(name)

        df = df_raw.iloc[header_row + 1 :].copy()
        df.columns = deduped
        df = df.reset_index(drop=True)

        # Normalize: replace pandas NaN-equivalent strings with actual None/NaN
        # Keep everything as str | NaN for consistency
        return df

    @staticmethod
    def _read_csv_raw(path: Path, delimiter: str, encoding: str) -> list[list[str]]:
        """Read a CSV file into a list of string lists using csv.reader.

        Handles ragged rows (different column counts) gracefully, which is
        common when metadata lines precede the actual header row.
        """
        rows: list[list[str]] = []
        with path.open(newline="", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row in reader:
                rows.append([cell if cell is not None else "" for cell in row])
        return rows

    @staticmethod
    def _detect_encoding(raw_bytes: bytes) -> str:
        sample = raw_bytes[:32768]
        result = chardet.detect(sample)
        encoding = result.get("encoding") or "utf-8"
        confidence = result.get("confidence", 0.0)
        # If confidence is low, prefer utf-8 as safe fallback
        if confidence < 0.7:
            encoding = "utf-8"
        # Normalise some common aliases first (handles chardet quirks like "UTF-8-BOM")
        normalized = encoding.lower().replace("-", "").replace("_", "")
        alias_map = {
            "utf8": "utf-8",
            "utf8bom": "utf-8-sig",
            "utf16": "utf-16",
            "latin1": "latin-1",
            "iso88591": "latin-1",
            "windows1252": "cp1252",
        }
        candidate = alias_map.get(normalized, encoding)
        # Use codecs.lookup() to get the canonical form Python actually uses
        try:
            candidate = codecs.lookup(candidate).name
        except LookupError:
            candidate = "utf-8"
        return candidate

    @staticmethod
    def _detect_delimiter(raw_bytes: bytes, encoding: str) -> str:
        sample = raw_bytes[:32768].decode(encoding, errors="replace")
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            return dialect.delimiter
        except csv.Error:
            # Count occurrences of each candidate
            counts = {d: sample.count(d) for d in [";", ",", "\t", "|"]}
            best = max(counts, key=lambda k: counts[k])
            return best if counts[best] > 0 else ","


def preview_header_rows(
    path: str | Path,
    n: int = 15,
    encoding_hint: str | None = None,
    delimiter_hint: str | None = None,
) -> list[list[str]]:
    """Return the first `n` raw rows of a CSV/XLSX file as lists of strings.

    Used by the load dialog to let the user pick the header row visually.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix in {".xlsx", ".xls", ".xlsm", ".ods"}:
        df_raw = pd.read_excel(
            path, header=None, nrows=n, dtype=str, engine="openpyxl"
        )
    else:
        raw_bytes = path.read_bytes()
        encoding = encoding_hint or DatasetLoader._detect_encoding(raw_bytes)
        delimiter = delimiter_hint or DatasetLoader._detect_delimiter(raw_bytes, encoding)
        # Use csv.reader so ragged rows (metadata lines) don't cause parse errors
        result: list[list[str]] = []
        with path.open(newline="", encoding=encoding, errors="replace") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader):
                if i >= n:
                    break
                result.append([cell if cell is not None else "" for cell in row])
        return result

    result: list[list[str]] = []
    for _, row in df_raw.iterrows():
        result.append([str(v) if pd.notna(v) else "" for v in row])
    return result


def get_xlsx_sheet_names(path: str | Path) -> list[str]:
    """Return the list of sheet names for an XLSX workbook."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names

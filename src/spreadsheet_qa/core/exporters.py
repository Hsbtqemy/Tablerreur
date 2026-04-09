"""Exporters: XLSX, CSV (always ;), TXT report, issues.csv, annotated exports."""

from __future__ import annotations

import csv
import io
import textwrap
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

from spreadsheet_qa.core.models import DatasetMeta, Issue, IssueStatus, Severity

# ---------------------------------------------------------------------------
# Lazy import of UI strings (core must remain Qt-free; t() is pure Python)
# ---------------------------------------------------------------------------


def _t(key: str, **kwargs: object) -> str:
    """Resolve a UI string key.  Falls back to key if i18n not available."""
    try:
        from spreadsheet_qa.ui.i18n import t
        return t(key, **kwargs)
    except Exception:
        return key


def _now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M")


_ANNOTATED_ROW_COL = "__tablerreur_ligne"
_ANNOTATED_STATUS_COL = "__tablerreur_statut"
_ANNOTATED_ISSUES_COL = "__tablerreur_anomalies"

_SEVERITY_RANK = {
    Severity.ERROR: 0,
    Severity.WARNING: 1,
    Severity.SUSPICION: 2,
}


def _issue_sort_key(issue: Issue) -> tuple[int, str, str]:
    return (
        _SEVERITY_RANK.get(issue.severity, 99),
        issue.col or "",
        issue.message or "",
    )


def _row_status_label(row_issues: list[Issue]) -> str:
    if not row_issues:
        return ""
    top_issue = min(row_issues, key=_issue_sort_key)
    return top_issue.severity.value


def _format_issue_line(issue: Issue) -> str:
    parts = [issue.severity.value]
    if issue.status != IssueStatus.OPEN:
        parts.append(issue.status.value)
    head = " / ".join(parts)
    return f"{head} — {issue.col}: {issue.message}"


def build_annotated_dataframe(
    df: pd.DataFrame,
    issues: list[Issue],
    row_positions: list[int] | None = None,
    include_status_column: bool = True,
) -> pd.DataFrame:
    """Return a DataFrame enriched with row-level Tablerreur annotations."""
    positions = list(row_positions) if row_positions is not None else list(range(len(df)))
    export_df = df.iloc[positions].copy()
    if not include_status_column:
        return export_df

    issues_by_row: dict[int, list[Issue]] = defaultdict(list)
    for issue in issues:
        issues_by_row[issue.row].append(issue)

    export_df.insert(0, _ANNOTATED_ROW_COL, [row_idx + 1 for row_idx in positions])
    export_df[_ANNOTATED_STATUS_COL] = [
        _row_status_label(sorted(issues_by_row.get(row_idx, []), key=_issue_sort_key))
        for row_idx in positions
    ]
    export_df[_ANNOTATED_ISSUES_COL] = [
        " | ".join(
            _format_issue_line(issue)
            for issue in sorted(issues_by_row.get(row_idx, []), key=_issue_sort_key)
        )
        for row_idx in positions
    ]
    return export_df


# ---------------------------------------------------------------------------
# XLSX export
# ---------------------------------------------------------------------------


class XLSXExporter:
    """Export cleaned DataFrame to XLSX."""

    def export(self, df: pd.DataFrame, path: Path) -> None:
        import openpyxl
        from openpyxl.styles import Font

        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _t("xlsx.sheet.data")

        # Header row
        header_font = Font(bold=True)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font

        # Data rows
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                cell_val = "" if pd.isna(val) else str(val)
                ws.cell(row=row_idx, column=col_idx, value=cell_val)

        wb.save(path)


class AnnotatedXLSXExporter:
    """Export a worksheet with row annotations and optional visual marks."""

    def export(
        self,
        df: pd.DataFrame,
        path: Path,
        issues: list[Issue],
        row_positions: list[int] | None = None,
        touched_cells: set[tuple[int, str]] | None = None,
        include_visual_marks: bool = True,
        include_status_column: bool = True,
    ) -> None:
        import openpyxl
        from openpyxl.comments import Comment
        from openpyxl.styles import Font, PatternFill

        positions = list(row_positions) if row_positions is not None else list(range(len(df)))
        touched = set(touched_cells or set())
        export_df = build_annotated_dataframe(
            df,
            issues,
            row_positions=positions,
            include_status_column=include_status_column,
        )

        path.parent.mkdir(parents=True, exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = _t("xlsx.sheet.data")

        header_font = Font(bold=True)
        for col_idx, col_name in enumerate(export_df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font

        for row_idx, row in enumerate(export_df.itertuples(index=False), start=2):
            for col_idx, val in enumerate(row, start=1):
                cell_val = "" if pd.isna(val) else str(val)
                ws.cell(row=row_idx, column=col_idx, value=cell_val)

        if include_visual_marks:
            fills = {
                Severity.ERROR: PatternFill(fill_type="solid", fgColor="FDE2E1"),
                Severity.WARNING: PatternFill(fill_type="solid", fgColor="FDF0D5"),
                Severity.SUSPICION: PatternFill(fill_type="solid", fgColor="E3E8FF"),
            }
            touched_fill = PatternFill(fill_type="solid", fgColor="DCEFFE")

            row_to_sheet_row = {row_idx: export_idx + 2 for export_idx, row_idx in enumerate(positions)}
            prefix_cols = 1 if include_status_column else 0
            source_col_to_sheet_col = {
                col_name: prefix_cols + offset
                for offset, col_name in enumerate(df.columns, start=1)
            }

            cell_issues: dict[tuple[int, str], list[Issue]] = defaultdict(list)
            row_issues: dict[int, list[Issue]] = defaultdict(list)
            for issue in issues:
                if issue.row not in row_to_sheet_row or issue.col not in source_col_to_sheet_col:
                    continue
                cell_issues[(issue.row, issue.col)].append(issue)
                row_issues[issue.row].append(issue)

            for (row_idx, col_name), grouped in cell_issues.items():
                cell = ws.cell(
                    row=row_to_sheet_row[row_idx],
                    column=source_col_to_sheet_col[col_name],
                )
                top_issue = min(grouped, key=_issue_sort_key)
                cell.fill = fills.get(top_issue.severity, cell.fill)
                cell.comment = Comment(
                    "\n".join(_format_issue_line(issue) for issue in sorted(grouped, key=_issue_sort_key)),
                    "Tablerreur",
                )

            if include_status_column:
                status_col_idx = export_df.columns.get_loc(_ANNOTATED_STATUS_COL) + 1
                for row_idx, grouped in row_issues.items():
                    top_issue = min(grouped, key=_issue_sort_key)
                    status_cell = ws.cell(row=row_to_sheet_row[row_idx], column=status_col_idx)
                    status_cell.fill = fills.get(top_issue.severity, status_cell.fill)

            for row_idx, col_name in touched:
                if (row_idx, col_name) in cell_issues:
                    continue
                sheet_row = row_to_sheet_row.get(row_idx)
                sheet_col = source_col_to_sheet_col.get(col_name)
                if sheet_row is None or sheet_col is None:
                    continue
                ws.cell(row=sheet_row, column=sheet_col).fill = touched_fill

        wb.save(path)


# ---------------------------------------------------------------------------
# CSV export (ALWAYS ; delimiter)
# ---------------------------------------------------------------------------


class CSVExporter:
    """Export cleaned DataFrame to CSV.

    Rules (non-negotiable):
    - Delimiter: ;
    - Quote char: "
    - Quoting: QUOTE_MINIMAL (cells with ; or " or newline are quoted)
    - Encoding: UTF-8 or UTF-8-BOM
    """

    def export(self, df: pd.DataFrame, path: Path, bom: bool = False) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        encoding = "utf-8-sig" if bom else "utf-8"

        with path.open("w", encoding=encoding, newline="") as f:
            writer = csv.writer(f, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(list(df.columns))
            for row in df.itertuples(index=False, name=None):
                writer.writerow(["" if pd.isna(v) else str(v) for v in row])


# ---------------------------------------------------------------------------
# Issues CSV export (ALWAYS ; delimiter)
# ---------------------------------------------------------------------------


class IssuesCSVExporter:
    """Export issues list to CSV with ; delimiter — French column headers."""

    @property
    def _columns(self) -> list[str]:
        return [
            _t("csv.issue_id"),
            _t("csv.severity"),
            _t("csv.status"),
            _t("csv.rule_id"),
            _t("csv.row"),
            _t("csv.column"),
            _t("csv.message"),
            _t("csv.original_value"),
            _t("csv.suggestion"),
            _t("csv.detected_at"),
        ]

    def export(self, issues: list[Issue], path: Path, meta: DatasetMeta | None = None) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        ts = _now_stamp()

        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter=";", quotechar='"', quoting=csv.QUOTE_MINIMAL)
            writer.writerow(self._columns)
            for issue in issues:
                writer.writerow([
                    issue.id,
                    issue.severity.value,
                    issue.status.value,
                    issue.rule_id,
                    issue.row + 1,  # 1-based for humans
                    issue.col,
                    issue.message,
                    str(issue.original) if issue.original is not None else "",
                    str(issue.suggestion) if issue.suggestion is not None else "",
                    ts,
                ])


# ---------------------------------------------------------------------------
# TXT report — French headings
# ---------------------------------------------------------------------------


class TXTReporter:
    """Generate a human-readable text report in French."""

    def export(
        self,
        issues: list[Issue],
        path: Path,
        meta: DatasetMeta | None = None,
        open_only: bool = True,
    ) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        lines: list[str] = []
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")

        # Header
        lines.append("=" * 72)
        lines.append(_t("report.header"))
        lines.append(f"{_t('report.generated')} {ts}")
        if meta:
            lines.append(f"{_t('report.source')} {meta.file_path}")
            lines.append(
                f"{_t('report.shape')} "
                f"{meta.original_shape[0]} lignes × {meta.original_shape[1]} colonnes"
            )
            lines.append(f"{_t('report.encoding')} {meta.encoding}")
        lines.append("=" * 72)
        lines.append("")

        filtered = [i for i in issues if i.status == IssueStatus.OPEN] if open_only else issues

        # Summary counts
        counts = Counter(i.severity for i in filtered)
        lines.append(_t("report.summary"))
        lines.append("-" * 40)
        severity_names = {
            Severity.ERROR: _t("severity.ERROR"),
            Severity.WARNING: _t("severity.WARNING"),
            Severity.SUSPICION: _t("severity.SUSPICION"),
        }
        for sev in [Severity.ERROR, Severity.WARNING, Severity.SUSPICION]:
            lines.append(f"  {severity_names[sev]:<16} {counts.get(sev, 0):>5}")
        lines.append(f"  {'TOTAL':<16} {len(filtered):>5}")
        lines.append("")

        # Top columns
        col_counts = Counter(i.col for i in filtered)
        if col_counts:
            lines.append(_t("report.top_cols"))
            lines.append("-" * 40)
            for col, cnt in col_counts.most_common(10):
                lines.append(f"  {col:<35} {cnt:>5} {_t('report.issue_count')}")
            lines.append("")

        # Top issue types
        type_counts = Counter(i.rule_id for i in filtered)
        if type_counts:
            lines.append(_t("report.top_types"))
            lines.append("-" * 40)
            for rule_id, cnt in type_counts.most_common(10):
                lines.append(f"  {rule_id:<45} {cnt:>5}")
            lines.append("")

        # Details (grouped by severity)
        lines.append(_t("report.details"))
        lines.append("=" * 72)
        for sev in [Severity.ERROR, Severity.WARNING, Severity.SUSPICION]:
            sev_issues = [i for i in filtered if i.severity == sev]
            if not sev_issues:
                continue
            sev_name = severity_names[sev]
            lines.append(f"\n[{sev_name}] — {len(sev_issues)} {_t('report.issue_count')}")
            lines.append("-" * 40)
            for issue in sev_issues[:200]:  # cap per severity
                loc = f"{_t('report.row')} {issue.row + 1}, «{issue.col}»"
                lines.append(f"  {loc}")
                lines.append(f"    {issue.message}")
                if issue.suggestion is not None:
                    lines.append(f"    {_t('report.suggestion')} {issue.suggestion!r}")
                lines.append("")

        path.write_text("\n".join(lines), encoding="utf-8")
